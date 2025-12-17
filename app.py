from flask import Flask, render_template, request
import joblib
import numpy as np
from openpyxl import Workbook, load_workbook
import os
import pandas as pd
import matplotlib.pyplot as plt

app = Flask(__name__)

# Load trained Random Forest model
model = joblib.load("rf_fraud_model.pkl")

# Excel file name
EXCEL_FILE = "history.xlsx"

# save transaction to excel
def save_history(type_, amount, risk, prediction):
    # Create Excel file if it doesn't exist
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        ws.append(["ID", "Type", "Amount", "Risk %", "Prediction"])
        wb.save(EXCEL_FILE)

    # Load workbook and append data
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    next_id = ws.max_row
    ws.append([next_id, type_, amount, risk, prediction])
    wb.save(EXCEL_FILE)

# -dashboard
@app.route("/", methods=["GET", "POST"])
def index():
    prediction = None
    prob = None

    if request.method == "POST":
        try:
            type_ = request.form["type"]
            amount = float(request.form["amount"])
            old_org = float(request.form["old_org"])
            new_org = float(request.form["new_org"])
            old_dest = float(request.form["old_dest"])
            new_dest = float(request.form["new_dest"])

            # Feature engineering
            error_org = old_org - new_org - amount
            error_dest = new_dest - old_dest - amount
            balance_diff_org = new_org - old_org
            balance_diff_dest = new_dest - old_dest
            ratio_org = new_org / (old_org + 1)
            ratio_dest = new_dest / (old_dest + 1)
            log_amount = np.log1p(amount)

            # One-hot encoding
            type_features = {"CASH_OUT": 0, "DEBIT": 0, "PAYMENT": 0, "TRANSFER": 0}
            type_features[type_] = 1

            # Feature vector
            features = [
                0, amount, old_org, new_org, old_dest, new_dest,
                error_org, error_dest, balance_diff_org, balance_diff_dest,
                ratio_org, ratio_dest,
                type_features["CASH_OUT"],
                type_features["DEBIT"],
                type_features["PAYMENT"],
                type_features["TRANSFER"],
                log_amount, 0, 0
            ]

            X_input = np.array([features])

            # Base ML Prediction
            base_prob = model.predict_proba(X_input)[0][1]
            risk = base_prob

            # Hybrid Rule-Based Enhancements
            if type_ in ["CASH_OUT", "TRANSFER", "DEBIT", "PAYMENT"]:
                if amount > 10000000:
                    risk += 0.25
                if new_org <= 0 and old_org > 0:
                    risk += 0.35
                if abs(new_dest - old_dest) < 1e-2:
                    risk += 0.40
                if amount > old_org:
                    risk += 0.45
                if abs((old_org - new_org) - amount) > 1e-2:
                    risk += 0.30
                if type_ == "TRANSFER" and abs((new_dest - old_dest) - amount) > 1e-2:
                    risk += 0.30
                if new_dest == 0 and old_dest == 0:
                    risk += 0.25

            # Cap & Probability
            risk = min(risk, 1.0)
            prob = round(risk * 100, 2)

            # Classification
            if risk > 0.7:
                prediction = "High Risk Fraud"
            elif risk > 0.3:
                prediction = "Suspicious Transaction"
            else:
                prediction = "Legitimate Transaction"

            # Save to Excel
            save_history(type_, amount, prob, prediction)

        except Exception as e:
            prediction = f"Error: {e}"

    return render_template("index.html", result=prediction, prob=prob)
#charts
@app.route("/charts")
def charts():
    if not os.path.exists(EXCEL_FILE):
        return "No data available yet."

    df = pd.read_excel(EXCEL_FILE)

    if df.empty:
        return "No transactions to visualize."
      # ---- Clean Prediction Column ----
    df["Prediction"] = df["Prediction"].str.strip()  # Remove leading/trailing spaces
    df["Prediction"] = df["Prediction"].replace({"High Risk Fraud": "High Risk Fraud",
                                                 "Suspicious Transaction": "Suspicious Transaction",
                                                 "Legitimate Transaction": "Legitimate Transaction"})  # Optional standardization

    # Chart 1: Prediction Count
    plt.figure()
    df["Prediction"].value_counts().plot(kind="bar", color="skyblue")
    plt.title("Transaction Prediction Distribution")
    plt.xlabel("Prediction Type")
    plt.ylabel("Count")
    plt.tight_layout()
    if not os.path.exists("static"):
        os.makedirs("static")
    plt.savefig("static/prediction_count.png")
    plt.close()

    # Chart 2: Risk Distribution
    plt.figure()
    plt.hist(df["Risk %"], bins=10, color="salmon")
    plt.title("Risk Percentage Distribution")
    plt.xlabel("Risk %")
    plt.ylabel("Frequency")
    plt.tight_layout()
    plt.savefig("static/risk_distribution.png")
    plt.close()

    return render_template("charts.html")

# ------------------- HISTORY PAGE -------------------
@app.route("/history")
def history():
    if not os.path.exists(EXCEL_FILE):
        return "No transaction history available."

    df = pd.read_excel(EXCEL_FILE)

    if df.empty:
        return "No transactions to display."

    # Convert dataframe to list of rows for Jinja2
    data = df.values.tolist()
    return render_template("history.html", data=data)

# ------------------- RUN APP -------------------
if __name__ == "__main__":
    app.run(debug=True, port=8000)
